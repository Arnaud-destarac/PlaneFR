"""
Ajoute a la feuille "synthese" de "facteurs de caracterisation.xlsx" une ligne par
colonne (extension exiobase) de "impact_world_plus_2.2.1_expert_version_exiobase_3.9_and_after.xlsx"
dont la somme des lignes en unite PDF.m2.yr (metrique biodiversite d'Impact World+) est non nulle.

Seules les lignes dont le libelle se termine par "(PDF.m2.yr)" sont sommees : ce sont les
indicateurs relatifs a la biodiversite (les autres lignes du fichier - DALY, CTUh, kg eq,
m3 eq, etc. - portent sur d'autres categories d'impact et ne doivent pas etre melangees).

Pour chaque colonne retenue :
    - "Extensions exiobase"          -> nom de la colonne
    - "Unite"                        -> laisse vide
    - "Sous-processus"               -> "Biodiversity loss"
    - "Processus du systeme Terre"   -> laisse vide
    - "Facteurs de caracterisation"  -> somme des valeurs (lignes PDF.m2.yr) de la colonne
"""

import shutil
from pathlib import Path

import openpyxl

DATA_DIR = Path(__file__).resolve().parent
SYNTHESE_PATH = DATA_DIR / "facteurs de caractérisation.xlsx"
IMPACT_WORLD_PATH = DATA_DIR / "impact_world_plus_2.2.1_expert_version_exiobase_3.9_and_after.xlsx"

SHEET_SYNTHESE = "synthèse"
SHEET_IMPACT_WORLD = "Sheet1"

COL_PROCESSUS_TERRE = 1  # A
COL_SOUS_PROCESSUS = 2  # B
COL_EXTENSIONS_EXIOBASE = 3  # C
COL_UNITE = 4  # D
COL_FACTEURS_CARACTERISATION = 5  # E

SOUS_PROCESSUS_VALUE = "Biodiversity loss"


PDF_UNIT_SUFFIX = "(PDF.m2.yr)"
EXCLUDED_SUBSTRING = "long term"


def get_biodiversity_rows(ws):
    """Return the row indices whose label is a PDF.m2.yr (biodiversity) indicator,
    excluding any indicator whose name contains "long term"."""
    rows = []
    for r in range(2, ws.max_row + 1):
        label = ws.cell(row=r, column=1).value
        if not isinstance(label, str):
            continue
        if not label.strip().endswith(PDF_UNIT_SUFFIX):
            continue
        if EXCLUDED_SUBSTRING in label.lower():
            continue
        rows.append(r)
    return rows


def compute_nonzero_column_sums(ws, biodiversity_rows):
    """Return list of (column_name, sum) for each data column whose sum over the
    biodiversity (PDF.m2.yr) rows is non-zero."""
    max_col = ws.max_column
    results = []
    for c in range(2, max_col + 1):  # column 1 is the row-label column
        colname = ws.cell(row=1, column=c).value
        total = 0.0
        for r in biodiversity_rows:
            v = ws.cell(row=r, column=c).value
            if isinstance(v, (int, float)):
                total += v
        if total != 0:
            results.append((colname, total))
    return results


def main():
    backup_path = SYNTHESE_PATH.with_name(SYNTHESE_PATH.stem + " (backup avant ajout biodiversity loss).xlsx")
    if backup_path.exists():
        # Repart toujours de la sauvegarde pristine pour eviter d'empiler des lignes
        # a chaque nouvelle execution du script (ex: apres correction du calcul).
        shutil.copy2(backup_path, SYNTHESE_PATH)
        print(f"Fichier restauré depuis la sauvegarde : {backup_path.name}")
    else:
        shutil.copy2(SYNTHESE_PATH, backup_path)
        print(f"Sauvegarde créée : {backup_path.name}")

    wb_impact = openpyxl.load_workbook(IMPACT_WORLD_PATH, data_only=True)
    ws_impact = wb_impact[SHEET_IMPACT_WORLD]
    biodiversity_rows = get_biodiversity_rows(ws_impact)
    print(f"{len(biodiversity_rows)} lignes '(PDF.m2.yr)' identifiées comme indicateurs de biodiversité.")
    nonzero_sums = compute_nonzero_column_sums(ws_impact, biodiversity_rows)
    print(f"{len(nonzero_sums)} colonnes avec somme non nulle sur {ws_impact.max_column - 1} colonnes.")

    wb_synthese = openpyxl.load_workbook(SYNTHESE_PATH, data_only=False)
    ws_synthese = wb_synthese[SHEET_SYNTHESE]

    next_row = ws_synthese.max_row + 1
    for offset, (colname, total) in enumerate(nonzero_sums):
        r = next_row + offset
        ws_synthese.cell(row=r, column=COL_PROCESSUS_TERRE, value=None)
        ws_synthese.cell(row=r, column=COL_SOUS_PROCESSUS, value=SOUS_PROCESSUS_VALUE)
        ws_synthese.cell(row=r, column=COL_EXTENSIONS_EXIOBASE, value=colname)
        ws_synthese.cell(row=r, column=COL_UNITE, value=None)
        ws_synthese.cell(row=r, column=COL_FACTEURS_CARACTERISATION, value=total)

    wb_synthese.save(SYNTHESE_PATH)
    print(f"{len(nonzero_sums)} lignes ajoutées à la feuille '{SHEET_SYNTHESE}' (lignes {next_row} à {next_row + len(nonzero_sums) - 1}).")


if __name__ == "__main__":
    main()
